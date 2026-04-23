import openpyxl
from openpyxl.utils import get_column_letter

# Caminho do arquivo Excel
EXCEL_PATH = '/Volumes/Expansion/----- MAMI -----/GESTAO_CURSOS_2026.xlsx'

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)

print('Abas disponíveis:')
for sheet in wb.sheetnames:
    print(f'- {sheet}')
    ws = wb[sheet]
    print('  Colunas:')
    for col in ws.iter_cols(min_row=1, max_row=1):
        col_name = col[0].value
        if col_name:
            print(f'    - {col_name}')
    print('  Fórmulas (primeiras 10 linhas):')
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.data_type == 'f':
                col_letter = get_column_letter(cell.column)
                print(f'    {col_letter}{cell.row}: {cell.value}')
    print()
