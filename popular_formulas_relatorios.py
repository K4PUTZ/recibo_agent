import openpyxl

EXCEL_PATH = '/Volumes/Expansion/----- MAMI -----/GESTAO_CURSOS_2026.xlsx'
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['RELATORIOS']

# Cursos e meses
cursos = ['CANTO CURAR', 'NINAR (KANTELE)', 'LIRA', 'PARTICULAR']
meses = ['JAN','FEV','MAR','ABR','MAI','JUN','JUL','AGO','SET','OUT','NOV','DEZ']

# Linha inicial dos cursos na tabela (ajuste se necessário)
linha_inicial = 12
coluna_inicial = 2  # B = 2

for i, curso in enumerate(cursos):
    for j, mes in enumerate(meses):
        # Exemplo de fórmula robusta para Excel 365/Online
        formula = (
            f'=SUMIFS(tblPagamentos[VALOR], '
            f'tblPagamentos[AlunoBeneficiario], '
            f'FILTER(tblMatriculas[StudentName], (tblMatriculas[CourseName]="{curso}")*(tblMatriculas[AnoLetivo]=YEAR(TODAY()))), '
            f'tblPagamentos[Competencia], "{mes}", '
            f'tblPagamentos[StatusPagamento], "Confirmado")'
        )
        cell = ws.cell(row=linha_inicial + i, column=coluna_inicial + j)
        cell.value = formula

# Fórmula de TOTAL por curso (soma dos meses)
for i in range(len(cursos)):
    start_col = openpyxl.utils.get_column_letter(coluna_inicial)
    end_col = openpyxl.utils.get_column_letter(coluna_inicial + len(meses) - 1)
    row = linha_inicial + i
    ws[f'O{row}'] = f'=SUM({start_col}{row}:{end_col}{row})'

# Fórmula de TOTAL por mês (soma dos cursos)
for j in range(len(meses)):
    col = openpyxl.utils.get_column_letter(coluna_inicial + j)
    ws[f'{col}{linha_inicial + len(cursos)}'] = f'=SUM({col}{linha_inicial}:{col}{linha_inicial + len(cursos) - 1})'

# Fórmula de TOTAL GERAL (soma dos totais dos cursos)
ws[f'O{linha_inicial + len(cursos)}'] = f'=SUM(O{linha_inicial}:O{linha_inicial + len(cursos) - 1})'

wb.save(EXCEL_PATH)
print('Fórmulas de RECEITA POR CURSO atualizadas!')
